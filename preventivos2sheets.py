#!/usr/bin/env python3
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
import time
import pandas as pd
import pygerduty.v2
import datetime
import requests
import sys
import os
import json
import openpyxl
import argparse

def as_text(value):
    if value is None:
        return ""
    return str(value)

def status_color(status):
    if (status == 'red'):
        return '00FF0000'
    elif (status == 'green'):
        return '0000FF00'
    elif (status == 'yellow'):
        return '00FFFF00'

def null_args(args):
    return [a for a in vars(args) if not getattr(args, a)]

parser = argparse.ArgumentParser(
    formatter_class=argparse.RawTextHelpFormatter,
    description="Prepara a planilha com dados dos preventivos do cliente.\n\n"+
                "Necessario abrir porta 80 para o IP range 18.235.142.155/32 no Zabbix Server.")
parser.add_argument("--akid", metavar="AWS_ACCESS_KEY_ID", default=os.environ.get('AWS_ACCESS_KEY_ID'),
                  help="Seu AWS Access Key ID da conta principal.")
parser.add_argument("--aksecret", metavar="AWS_SECRET_ACCESS_KEY", default=os.environ.get('AWS_SECRET_ACCESS_KEY'),
                  help="Seu AWS Secret Key da conta principal.")
parser.add_argument("--alias", metavar="SWITCH_ROLE_ALIAS", default=os.environ.get('SWITCH_ROLE_ALIAS'),
                  help="Alias da conta de switch-role do cliente.")
parser.add_argument("--zurl", metavar="ZABBIX_USER_URL", default=os.environ.get('ZABBIX_USER_URL'),
                  help="A URL do Zabbix Server do cliente.")
parser.add_argument("--zuser", metavar="ZABBIX_SERVER_USER", default=os.environ.get('ZABBIX_SERVER_USER'),
                  help="O usuario do Zabbix Server do cliente.")
parser.add_argument("--zpass", metavar="ZABBIX_SERVER_PASSWORD", default=os.environ.get('ZABBIX_SERVER_PASSWORD'),
                  help="A senha do usuario do Zabbix Server do cliente.")
parser.add_argument("--service", metavar="PAGERDUTY_SERVICE", default=os.environ.get('PAGERDUTY_SERVICE'),
                  help="O serviço do cliente no Pager Duty.")
parser.add_argument("--pduser", metavar="PAGERDUTY_USER_EMAIL", default=os.environ.get('PAGERDUTY_USER_EMAIL'),
                  help="O email de usuario do Pager Duty.")
parser.add_argument("--pdpass", metavar="PAGERDUTY_USER_PASSWORD", default=os.environ.get('PAGERDUTY_USER_PASSWORD'),
                  help="A senha de usuario do Pager Duty.")
parser.add_argument("--pdkey", metavar="PAGERDUTY_API_KEY", default=os.environ.get('PAGERDUTY_API_KEY'),
                    help="A API Key do PagerDuty.")
parser.add_argument("--out", metavar="OUTPUT_FILE_NAME", default='preventivos2sheets',
                    help="Nome do arquivo de output do sheets.")

required = ['akid', 'aksecret', 'zurl', 'zuser', 'zpass', 'service', 'pduser', 'pdpass', 'pdkey']
args = parser.parse_args()
empty_args = null_args(args)

for a in required:
    if a in empty_args:
        parser.print_help()
        print('\nthe following arguments are required: --{}'.format(', --'.join(empty_args)))
        exit()

h = {'akid': args.akid,
     'aksecret': args.aksecret}

if args.alias:
    h['alias'] = args.alias

wb = openpyxl.Workbook()
del wb['Sheet']
sheets = []
print(' ')

# PREVENTIVO PAGERDUTY
print('Inciando revisão do PagerDuty...')
profile = webdriver.FirefoxProfile()
profile.set_preference("browser.download.folderList", 2)
profile.set_preference("browser.download.manager.showWhenStarting", False)
profile.set_preference("browser.download.dir", '/tmp')
profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv")
options = Options()
options.headless = True
print('Iniciando driver...')
with webdriver.Firefox(firefox_profile=profile, options=options, log_path="/tmp/geckodriver.log") as br:
    if os.path.exists("/tmp/incidents.csv"):
        print('Removendo /tmp/incidents.csv existente...')
        os.remove("/tmp/incidents.csv")
    print('Conectando ao PagerDuty...')
    br.get('https://rivendel-tecnologia.pagerduty.com/sign_in')
    print('Acessando...')
    br.find_element_by_id('user_email').send_keys(args.pduser)
    br.find_element_by_id('user_password').send_keys(args.pdpass)
    br.find_element_by_name('commit').click()
    print('Localizando serviço...')
    pager = pygerduty.v2.PagerDuty(args.pdkey)
    service = list(pager.services.list(query=args.service))[0]
    since = (datetime.datetime.now() + datetime.timedelta(-30))
    since = since.strftime("%Y-%m-%dT%H:%M:%S").replace(':', '%3A')
    until = datetime.datetime.now().strftime(
        "%Y-%m-%dT%H:%M:%S").replace(':', '%3A')
    params = 'since={}&until={}&filters[service_ids][]={}&time_zone=America%2FSao_Paulo'.format(
        since, until, service.id)

    print('Carregando incidentes...')
    br.get('https://rivendel-tecnologia.pagerduty.com/reports#drilldown?{}'.format(params))
    br.find_element_by_class_name('reports-csv-download').click()
    time.sleep(3)
    while os.path.isfile('/tmp/incidents.csv.part'):
        time.sleep(1)
    print('Incidentes carregados!')
    br.close()

df = pd.read_csv('/tmp/incidents.csv')

wb.create_sheet('PagerDuty')
pd = wb['PagerDuty']
pd.cell(row=1, column=1).value = 'PagerDuty SLA'
pd.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
pd.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
pd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
pd.cell(row=2, column=1).value = 'MTTA'
pd.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
pd.cell(row=2, column=2).value = 'MTTR'
pd.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
pd.cell(row=3, column=1).value = df['seconds_to_first_ack'].mean()
pd.cell(row=3, column=2).value = df['seconds_to_resolve'].mean()

pd.cell(row=5, column=1).value = 'PagerDuty Incidents'
pd.cell(row=5, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
pd.cell(row=5, column=1).font = openpyxl.styles.Font(bold=True)
pd.merge_cells(start_row=5, start_column=1, end_row=5, end_column=23)
pd.cell(row=6, column=1).value = 'Id'
pd.cell(row=6, column=1).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=2).value = 'Incident Number'
pd.cell(row=6, column=2).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=3).value = 'Description'
pd.cell(row=6, column=3).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=4).value = 'Service Id'
pd.cell(row=6, column=4).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=5).value = 'Service Name'
pd.cell(row=6, column=5).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=6).value = 'Escalation Policy Id'
pd.cell(row=6, column=6).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=7).value = 'Escalation Policy Name'
pd.cell(row=6, column=7).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=8).value = 'Created On'
pd.cell(row=6, column=8).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=9).value = 'Resolved On'
pd.cell(row=6, column=9).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=10).value = 'Seconds to First Ack'
pd.cell(row=6, column=10).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=11).value = 'Seconds to Resolve'
pd.cell(row=6, column=11).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=12).value = 'Auto Resolved'
pd.cell(row=6, column=12).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=13).value = 'Escalation Count'
pd.cell(row=6, column=13).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=14).value = 'Auto Escalation Count'
pd.cell(row=6, column=14).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=15).value = 'Acknowledge Count'
pd.cell(row=6, column=15).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=16).value = 'Assignment Count'
pd.cell(row=6, column=16).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=17).value = 'Acknowledged by User Ids'
pd.cell(row=6, column=17).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=18).value = 'Acknowledged by User Names'
pd.cell(row=6, column=18).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=19).value = 'Assigned by User Ids'
pd.cell(row=6, column=19).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=20).value = 'Assigned by User Names'
pd.cell(row=6, column=20).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=21).value = 'Resolved by User Ids'
pd.cell(row=6, column=21).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=22).value = 'Resolved by User Names'
pd.cell(row=6, column=22).font = openpyxl.styles.Font(bold=True)
pd.cell(row=6, column=23).value = 'Urgency'
pd.cell(row=6, column=23).font = openpyxl.styles.Font(bold=True)

r = 7
for index, row in df.iterrows():
    pd.cell(row=r, column=1).value = row['id']
    pd.cell(row=r, column=2).value = row['incident_number']
    pd.cell(row=r, column=3).value = row['description']
    pd.cell(row=r, column=4).value = row['service_id']
    pd.cell(row=r, column=5).value = row['service_name']
    pd.cell(row=r, column=6).value = row['escalation_policy_id']
    pd.cell(row=r, column=7).value = row['escalation_policy_name']
    pd.cell(row=r, column=8).value = row['created_on']
    pd.cell(row=r, column=9).value = row['resolved_on']
    pd.cell(row=r, column=10).value = row['seconds_to_first_ack']
    pd.cell(row=r, column=11).value = row['seconds_to_resolve']
    pd.cell(row=r, column=12).value = row['auto_resolved']
    pd.cell(row=r, column=13).value = row['escalation_count']
    pd.cell(row=r, column=14).value = row['auto_escalation_count']
    pd.cell(row=r, column=15).value = row['acknowledge_count']
    pd.cell(row=r, column=16).value = row['assignment_count']
    pd.cell(row=r, column=17).value = row['acknowledged_by_user_ids']
    pd.cell(row=r, column=18).value = row['acknowledged_by_user_names']
    pd.cell(row=r, column=19).value = row['assigned_to_user_ids']
    pd.cell(row=r, column=20).value = row['assigned_to_user_names']
    pd.cell(row=r, column=21).value = row['resolved_by_user_id']
    pd.cell(row=r, column=22).value = row['resolved_by_user_name']
    pd.cell(row=r, column=23).value = row['urgency']
    r = r+1
print('Revisão do PagerDuty completa!\n')
sheets.append(pd)
# FIM PREVENTIVO PAGERDUTY

# PREVENTIVO ZABBIX
print('Inciando revisão do Zabbix...')
h['zurl'] = args.zurl
h['zuser'] = args.zuser
h['zpass'] = args.zpass
q = requests.post(
    'https://api.rivendel.com.br/v1/preventivo/zabbix/compare', headers=h).json()

wb.create_sheet('Zabbix')
zabb = wb['Zabbix']
zabb.cell(row=1, column=1).value = 'Monitoring'
zabb.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
zabb.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
zabb.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
zabb.cell(row=2, column=1).value = 'Host Name'
zabb.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=2, column=2).value = 'Instance Name'
zabb.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=2, column=3).value = 'Instance Id'
zabb.cell(row=2, column=3).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=2, column=4).value = 'Common Ip'
zabb.cell(row=2, column=4).font = openpyxl.styles.Font(bold=True)

r = 3
for host in q['matches']:
    zabb.cell(row=r, column=1).value = host['ZabbixHost']
    zabb.cell(row=r, column=2).value = host['InstanceName']
    zabb.cell(row=r, column=3).value = host['InstanceId']
    zabb.cell(row=r, column=4).value = host['CommonIp']
    r = r+1

r = r+1
zabb.cell(row=r, column=1).value = 'Neglected Instances'
zabb.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
zabb.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
zabb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r = r+1
zabb.cell(row=r, column=1).value = 'Instance Name'
zabb.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=r, column=2).value = 'Instance Id'
zabb.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=r, column=3).value = 'Private Ip'
zabb.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=r, column=4).value = 'Public Ip'
zabb.cell(row=r, column=4).font = openpyxl.styles.Font(bold=True)

r = r+1
for inst in q['orphanedInstances']:
    zabb.cell(row=r, column=1).value = inst['name']
    zabb.cell(row=r, column=2).value = inst['id']
    zabb.cell(row=r, column=3).value = inst['privateIp']
    zabb.cell(row=r, column=4).value = inst['publicIp']
    r = r+1

r = r+1
zabb.cell(row=r, column=1).value = 'Forgotten Hosts'
zabb.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
zabb.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
zabb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r = r+1
zabb.cell(row=r, column=1).value = 'Host Name'
zabb.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
zabb.cell(row=r, column=2).value = 'Registered Ip'
zabb.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)

r = r+1
for inst in q['orphanedHosts']:
    zabb.cell(row=r, column=1).value = inst['name']
    zabb.cell(row=r, column=2).value = inst['ip']
    r = r+1

sheets.append(zabb)
del h['zurl']
del h['zuser']
del h['zpass']
print('Revisão do Zabbix completa!\n')
# FIM PREVENTIVO ZABBIX

# PREVENTIVO IAM
print('Inciando revisão de IAM...')
q = requests.post('https://api.rivendel.com.br/v1/preventivo/iam', headers=h).json()

wb.create_sheet('IAM')
iam = wb['IAM']
iam.cell(row=1, column=1).value = 'IAM'
iam.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
iam.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
iam.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
iam.cell(row=2, column=2).value = 'Actions'
iam.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
iam.cell(row=3, column=1).value = 'Main Account'
iam.cell(row=3, column=1).font = openpyxl.styles.Font(bold=True)
iam.cell(row=3, column=2).value = ';\n'.join(q['Account'])
iam.cell(row=4, column=1).value = 'Status'
iam.cell(row=4, column=1).font = openpyxl.styles.Font(bold=True)
iam.cell(row=4, column=2).value = q['Status']
iam.cell(row=4, column=2).font = openpyxl.styles.Font(color=status_color(q['Status'].lower()))
iam.cell(row=6, column=1).value = 'Users'
iam.cell(row=6, column=1).font = openpyxl.styles.Font(bold=True)
iam.cell(row=6, column=2).value = 'Actions'
iam.cell(row=6, column=2).font = openpyxl.styles.Font(bold=True)

r=7
for u in q['Users']:
    iam.cell(row=r, column=1).value = u
    iam.cell(row=r, column=2).value = ';\n'.join(q['Users'][u]['Actions'])
    r = r+1

sheets.append(iam)
print('Revisão de IAM completa!\n')
# FIM PREVENTIVO IAM

# PREVENTIVO SG
print('Inciando revisão de Security Groups...')
q = requests.post('https://api.rivendel.com.br/v1/preventivo/sg', headers=h).json()

wb.create_sheet('Security Groups')
sg = wb['Security Groups']
sg.cell(row=1, column=1).value = 'Unsafe Groups'
sg.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
sg.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
sg.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
sg.cell(row=2, column=1).value = 'Description'
sg.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
sg.cell(row=2, column=2).value = 'Group Id'
sg.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
sg.cell(row=2, column=3).value = 'Group Name'
sg.cell(row=2, column=3).font = openpyxl.styles.Font(bold=True)
sg.cell(row=2, column=4).value = 'Unsafe Ports'
sg.cell(row=2, column=4).font = openpyxl.styles.Font(bold=True)

r = 3
for unsafe_sg in q['SecurityGroups']['UnsafeGroups']:
    sg.cell(row=r, column=1).value = unsafe_sg['Description']
    sg.cell(row=r, column=2).value = unsafe_sg['GroupId']
    sg.cell(row=r, column=3).value = unsafe_sg['GroupName']
    sg.cell(row=r, column=4).value = ';\n'.join(
        ['{}:{}:{} -> {}'.format(p['IpProtocol'], p['CidrIp'], str(p.get('FromPort')), str(p.get('ToPort'))) for p in unsafe_sg['UnsafePorts']])
    r = r+1

r = r+1
sg.cell(row=r, column=1).value = 'Unused Groups'
sg.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
sg.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
sg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r=r+1
sg.cell(row=r, column=1).value = 'Description'
sg.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
sg.cell(row=r, column=2).value = 'Group Id'
sg.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)
sg.cell(row=r, column=3).value = 'Group Name'
sg.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)
sg.cell(row=r, column=4).value = 'Vpc Id'
sg.cell(row=r, column=4).font = openpyxl.styles.Font(bold=True)

r=r+1
for unused_sg in q['SecurityGroups']['UnusedGroups']:
    sg.cell(row=r, column=1).value = unused_sg['Description']
    sg.cell(row=r, column=2).value = unused_sg['GroupId']
    sg.cell(row=r, column=3).value = unused_sg['GroupName']
    sg.cell(row=r, column=4).value = unused_sg['VpcId']
    r = r+1

sheets.append(sg)
print('Revisão de Security Groups completa!\n')
# FIM PREVENTIVO SG

# PREVENTIVO TAGS
print('Inciando revisão de Tags...')
h['tags'] = 'backup,_liga,_desliga'
q = requests.post('https://api.rivendel.com.br/v1/preventivo/tags', headers=h).json()

wb.create_sheet('Tags')
tags = wb['Tags']
tags.cell(row=1, column=1).value = 'Found'
tags.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
tags.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
tags.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
tags.cell(row=2, column=1).value = 'Instance Name'
tags.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
tags.cell(row=2, column=2).value = 'Tag'
tags.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
tags.cell(row=2, column=3).value = 'Value'
tags.cell(row=2, column=3).font = openpyxl.styles.Font(bold=True)

r=3
for t in q['found']:
    tags.cell(row=r, column=1).value = t['Name']
    tags.cell(row=r, column=2).value = t['Tag']
    tags.cell(row=r, column=3).value = t['Value']
    r = r+1

r=r+1
tags.cell(row=r, column=1).value = 'Missing'
tags.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
tags.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
tags.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r=r+1
tags.cell(row=r, column=1).value = 'Instance Name'
tags.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
tags.cell(row=r, column=2).value = 'Tag'
tags.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)

r=r+1
for t in q['missing']:
    tags.cell(row=r, column=1).value = t['Name']
    tags.cell(row=r, column=2).value = t['Tag']
    r = r+1

r = r+1
tags.cell(row=r, column=1).value = 'Empty'
tags.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
tags.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
tags.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r = r+1
tags.cell(row=r, column=1).value = 'Instance Id'
tags.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)

r = r+1
for t in q['empty']:
    tags.cell(row=r, column=1).value = t['Id']
    
    r = r+1

sheets.append(tags)
del h['tags']
print('Revisão de Tags completa!\n')
# FIM PREVENTIVO TAGS

# PREVENTIVO ELASTICIP
print('Inciando revisão de Elastic Ip...')
q = requests.post('https://api.rivendel.com.br/v1/preventivo/elasticip', headers=h).json()

wb.create_sheet('Elastic Ip')
eip = wb['Elastic Ip']
eip.cell(row=1, column=1).value = 'Elastic Ip'
eip.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
eip.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
eip.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

r=3
for region in q:
    if q[region]:
        eip.cell(row=r, column=1).value = region
        eip.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        eip.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        eip.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r=r+1
        eip.cell(row=r, column=1).value = 'Allocation Id'
        eip.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        eip.cell(row=r, column=2).value = 'Public Ip'
        eip.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)

        r=r+1
        for ip in q[region]:
            eip.cell(row=r, column=1).value = ip['AllocationId']
            eip.cell(row=r, column=2).value = ip['PublicIp']
            r = r+1
        r=r+1

sheets.append(eip)
print('Revisão de Elastic Ip completa!\n')
# FIM PREVENTIVO ELASTICIP

# PREVENTIVO VOLUMES
print('Inciando revisão de Volumes...')
q = requests.post(
    'https://api.rivendel.com.br/v1/preventivo/volumes', headers=h).json()

wb.create_sheet('Volumes')
vol = wb['Volumes']
vol.cell(row=1, column=1).value = 'Volumes'
vol.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
vol.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
vol.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

r = 3
for region in q:
    if q[region]:
        vol.cell(row=r, column=1).value = region
        vol.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center")
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r = r+1
        vol.cell(row=r, column=1).value = 'Volume Id'
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=2).value = 'Volume Type'
        vol.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=3).value = 'Volume Size'
        vol.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)

        r = r+1
        for v in q[region]:
            vol.cell(row=r, column=1).value = v['VolumeId']
            vol.cell(row=r, column=2).value = v['VolumeType']
            vol.cell(row=r, column=3).value = v['Size']
            r = r+1
        r = r+1

sheets.append(vol)
print('Revisão de Volumes completa!\n')
# FIM PREVENTIVO VOLUMES

# PREVENTIVO RESERVAS
print('Inciando revisão de Reservas...')
q = requests.post('https://api.rivendel.com.br/v1/preventivo/reservas', headers=h).json()

wb.create_sheet('Reservas')
vol = wb['Reservas']
vol.cell(row=1, column=1).value = 'Reservas'
vol.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
vol.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
vol.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

r = 3
for region in q:
    if q[region]:
        vol.cell(row=r, column=1).value = region
        vol.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center")
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r = r+1
        vol.cell(row=r, column=1).value = 'Date'
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=2).value = 'Instance Type'
        vol.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=3).value = 'Status'
        vol.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)

        r = r+1
        for v in q[region]:
            vol.cell(row=r, column=1).value = v['Date']
            vol.cell(row=r, column=2).value = v['InstanceType']
            vol.cell(row=r, column=3).value = v['Status']
            r = r+1
        r = r+1

sheets.append(vol)
print('Revisão de Reservas completa!\n')
# FIM PREVENTIVO RESERVAS

# PREVENTIVO VPN
print('Inciando revisão de VPN...')
q = requests.post(
    'https://api.rivendel.com.br/v1/preventivo/vpn', headers=h).json()
qq = requests.post(
    'https://api.rivendel.com.br/v1/preventivo/vpn/tunel', headers=h).json()

wb.create_sheet('VPN')
vol = wb['VPN']
vol.cell(row=1, column=1).value = 'VPN'
vol.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
vol.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
vol.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

r = 3
for region in q:
    if q[region]:
        vol.cell(row=r, column=1).value = region
        vol.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center")
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r = r+1
        vol.cell(row=r, column=1).value = 'Vpn Id'
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=2).value = 'Vpn Name'
        vol.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=3).value = 'Status'
        vol.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=4).value = 'In Use'
        vol.cell(row=r, column=4).font = openpyxl.styles.Font(bold=True)

        r = r+1
        for v in q[region]:
            for vv in qq[region]:
                if (vv['VpnId'] == v['VpnId']):
                    v.update(vv)
            vol.cell(row=r, column=1).value = v['VpnId']
            vol.cell(row=r, column=2).value = v['VpnName']
            vol.cell(row=r, column=3).value = v['Status']
            vol.cell(row=r, column=4).value = v['InUse']
            r = r+1
        r = r+1

sheets.append(vol)
print('Revisão de VPN completa!\n')
# FIM PREVENTIVO VPN

# PREVENTIVO IGW
print('Inciando revisão de Internet Gateway...')
q = requests.post(
    'https://api.rivendel.com.br/v1/preventivo/igw', headers=h).json()

wb.create_sheet('IGW')
vol = wb['IGW']
vol.cell(row=1, column=1).value = 'Internet Gateway'
vol.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
vol.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
vol.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

r = 3
for region in q:
    if q[region]:
        vol.cell(row=r, column=1).value = region
        vol.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center")
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r = r+1
        vol.cell(row=r, column=1).value = 'Igw Id'
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=2).value = 'In Use'
        vol.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)

        r = r+1
        for v in q[region]:
            vol.cell(row=r, column=1).value = v['IgwId']
            vol.cell(row=r, column=2).value = v['InUse']
            r = r+1
        r = r+1

sheets.append(vol)
print('Revisão de Internet Gateway completa!\n')
# FIM PREVENTIVO IGW

# PREVENTIVO CPU
print('Inciando revisão de Creditos de CPU...')
q = requests.post('https://api.rivendel.com.br/v1/preventivo/cpu', headers=h).json()

wb.create_sheet('CPU')
vol = wb['CPU']
vol.cell(row=1, column=1).value = 'CPU Credits'
vol.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center")
vol.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
vol.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

r = 3
for region in q:
    if q[region]:
        vol.cell(row=r, column=1).value = region
        vol.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center")
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r = r+1
        vol.cell(row=r, column=1).value = 'Instance Id'
        vol.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=2).value = 'Instance Type'
        vol.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=3).value = 'Available Credits'
        vol.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=4).value = 'Used Credits'
        vol.cell(row=r, column=4).font = openpyxl.styles.Font(bold=True)
        vol.cell(row=r, column=5).value = 'Used Percentage'
        vol.cell(row=r, column=5).font = openpyxl.styles.Font(bold=True)

        r = r+1
        for v in q[region]:
            vol.cell(row=r, column=1).value = v['InstanceId']
            vol.cell(row=r, column=2).value = v['InstanceType']
            vol.cell(row=r, column=3).value = v['AvailableCredits']
            vol.cell(row=r, column=4).value = v['UsedCredits']
            vol.cell(row=r, column=5).value = v['UsedPercentage']
            r = r+1
        r = r+1

sheets.append(vol)
print('Revisão de Creditos de CPU completa!\n')
# FIM PREVENTIVO CPU

# FORMATACAO
for ws in sheets:
    for column_cells in ws.columns:
        length = max(len(line) for cell in column_cells for line in as_text(cell.value).split('\n'))
        ws.column_dimensions[column_cells[0].column].width = length

print('Relatorio salvo como {}.xlsx\n'.format(args.out))
wb.save('{}.xlsx'.format(args.out))
                    

